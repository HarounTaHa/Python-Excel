import openpyxl

from pathlib import Path

path = Path('D:\_Python_Projects\Python-Excel', 'example.xlsx')

excel_file = openpyxl.load_workbook(path)

print(excel_file.sheetnames)
sheet1 = excel_file['ورقة1']
print(sheet1.title)
activeSheet = excel_file.active
print(activeSheet.title)

print(sheet1['A1'].value)
print(sheet1['B1'].value)
print(sheet1['C1'].value)
print(sheet1['C1'].row)
print(sheet1['C1'].column)
print(sheet1['C1'].coordinate)

print(sheet1.cell(row=1, column=3).value)

for i in range(1, 7):
    print(i, sheet1.cell(row=i, column=1).value)

print('-' * 50)
total = 0
for i in range(1, sheet1.max_row):
    print(sheet1.cell(row=i, column=1).value, sheet1.cell(row=i, column=2).value)
    total += sheet1.cell(row=i, column=2).value

print(f'The total salary of the employee is {total}$')

print('-' * 50)
print(sheet1.max_row)
print(sheet1.max_column)
