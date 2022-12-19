from pathlib import Path

import openpyxl

# create excel file
excel_file = openpyxl.Workbook()
print(excel_file.sheetnames)

# change sheet name
excel_sheet = excel_file.active
excel_sheet.title = "first_sheet"

# create sheet
excel_file.create_sheet()
excel_file.create_sheet()
excel_file.create_sheet(index=1, title='second_sheet')

# delete sheet
del excel_file['Sheet']

# write to sheet
sheet = excel_file['second_sheet']
sheet['A1'] = 'Hello, world'
print(sheet['A1'].value)

# write to excel practice
names = ['ahmed', 'haroun', 'ali']
first_sheet = excel_file['first_sheet']
for i in range(1, len(names) + 1):
    first_sheet.cell(row=i, column=3).value = names[i-1]

# save excel file
# path = Path('D:\_Python_Projects\Python-Excel', 'new.xlsx')
# excel_file.save(filename=path)
excel_file.save(filename="new.xlsx")
